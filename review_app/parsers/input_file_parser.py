"""Review App - User Input File Parser.

This is SEPARATE from RAG App's parsers.
Used for parsing user input files for validation.
"""
from pathlib import Path
from typing import Optional
import csv
import json

SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg", ".avif", ".tif", ".tiff"}


class UserInputFileParser:
    """Parses user input files (invoices, forms, etc.) for validation."""
    
    async def parse(self, file_path: Path) -> Optional[dict]:
        """Parse user input file."""
        suffix = file_path.suffix.lower()
        
        if suffix == ".pdf":
            return await self._parse_pdf(file_path)
        elif suffix == ".docx":
            return await self._parse_docx(file_path)
        elif suffix == ".csv":
            return await self._parse_csv(file_path)
        elif suffix == ".json":
            return await self._parse_json(file_path)
        elif suffix in [".txt", ".md"]:
            return await self._parse_text(file_path)
        elif suffix in SPREADSHEET_SUFFIXES:
            return await self._parse_spreadsheet(file_path)
        elif suffix in IMAGE_SUFFIXES:
            return {
                "file_name": file_path.name,
                "file_type": suffix.lstrip("."),
                "content": (
                    f"[Image attached: {file_path.name}. The UI can preview this image, "
                    "but text-only review currently cannot extract visual details automatically.]"
                ),
                "metadata": {"size_bytes": file_path.stat().st_size, "image": True},
            }
        else:
            return {
                "file_name": file_path.name,
                "file_type": suffix,
                "content": f"[Unsupported: {suffix}]",
                "metadata": {"error": "unsupported"}
            }
    
    async def _parse_pdf(self, file_path: Path) -> dict:
        try:
            import pypdf
            text_parts = []
            
            with open(file_path, 'rb') as f:
                reader = pypdf.PdfReader(f)
                for page_num, page in enumerate(reader.pages, 1):
                    text = page.extract_text()
                    if text:
                        text_parts.append(f"--- Page {page_num} ---\n{text}")
            
            return {
                "file_name": file_path.name,
                "file_type": "pdf",
                "content": "\n\n".join(text_parts),
                "metadata": {"pages": len(reader.pages)}
            }
        except Exception as e:
            return {
                "file_name": file_path.name,
                "file_type": "pdf",
                "content": f"[Error: {str(e)}]",
                "metadata": {"error": str(e)}
            }
    
    async def _parse_docx(self, file_path: Path) -> dict:
        try:
            import docx
            doc = docx.Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            
            return {
                "file_name": file_path.name,
                "file_type": "docx",
                "content": "\n\n".join(paragraphs),
                "metadata": {"paragraphs": len(paragraphs)}
            }
        except Exception as e:
            return {
                "file_name": file_path.name,
                "file_type": "docx",
                "content": f"[Error: {str(e)}]",
                "metadata": {"error": str(e)}
            }
    
    async def _parse_text(self, file_path: Path) -> dict:
        try:
            content = file_path.read_text(encoding='utf-8')
            return {
                "file_name": file_path.name,
                "file_type": file_path.suffix.lstrip('.'),
                "content": content,
                "metadata": {"size_bytes": file_path.stat().st_size}
            }
        except UnicodeDecodeError:
            content = file_path.read_text(encoding='latin-1')
            return {
                "file_name": file_path.name,
                "file_type": file_path.suffix.lstrip('.'),
                "content": content,
                "metadata": {"encoding": "latin-1"}
            }

    async def _parse_csv(self, file_path: Path) -> dict:
        try:
            with open(file_path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.DictReader(f))
        except UnicodeDecodeError:
            with open(file_path, newline="", encoding="latin-1") as f:
                rows = list(csv.DictReader(f))

        if not rows:
            return await self._parse_text(file_path)

        text_parts = []
        for row_num, row in enumerate(rows, 1):
            values = [f"{key}: {value}" for key, value in row.items() if value not in (None, "")]
            text_parts.append(f"Row {row_num}\n" + "\n".join(values))

        return {
            "file_name": file_path.name,
            "file_type": "csv",
            "content": "\n\n".join(text_parts),
            "metadata": {
                "rows": len(rows),
                "columns": list(rows[0].keys()) if rows else [],
                "size_bytes": file_path.stat().st_size,
                "records": rows,
            },
        }

    async def _parse_json(self, file_path: Path) -> dict:
        try:
            raw_text = file_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            raw_text = file_path.read_text(encoding="latin-1")

        try:
            data = json.loads(raw_text)
        except json.JSONDecodeError as exc:
            return {
                "file_name": file_path.name,
                "file_type": "json",
                "content": f"[JSON parse error: {exc}]",
                "metadata": {"error": str(exc)},
            }

        return {
            "file_name": file_path.name,
            "file_type": "json",
            "content": json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True),
            "metadata": {
                "top_level_type": type(data).__name__,
                "size_bytes": file_path.stat().st_size,
                "records": data if isinstance(data, list) else [data] if isinstance(data, dict) else [],
            },
        }

    async def _parse_spreadsheet(self, file_path: Path) -> dict:
        suffix = file_path.suffix.lower()
        records = []
        text_parts = []

        if suffix == ".xls":
            try:
                import xlrd
                workbook = xlrd.open_workbook(str(file_path))
                for sheet in workbook.sheets():
                    headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)] if sheet.nrows else []
                    for row_index in range(1 if headers else 0, sheet.nrows):
                        row = {
                            headers[col] if col < len(headers) and headers[col] else f"column_{col + 1}": sheet.cell_value(row_index, col)
                            for col in range(sheet.ncols)
                        }
                        records.append(row)
                        values = [f"{key}: {value}" for key, value in row.items() if value not in (None, "")]
                        text_parts.append(f"Sheet {sheet.name} Row {row_index + 1}\n" + "\n".join(values))
            except Exception as exc:
                return {
                    "file_name": file_path.name,
                    "file_type": suffix.lstrip("."),
                    "content": f"[Spreadsheet parse error: {exc}]",
                    "metadata": {"error": str(exc)},
                }
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(str(file_path), data_only=True, read_only=True)
                for sheet in workbook.worksheets:
                    rows = list(sheet.iter_rows(values_only=True))
                    headers = [str(value).strip() if value is not None else "" for value in rows[0]] if rows else []
                    for index, row_values in enumerate(rows[1 if headers else 0:], 1 if headers else 0):
                        row = {
                            headers[col] if col < len(headers) and headers[col] else f"column_{col + 1}": value
                            for col, value in enumerate(row_values)
                        }
                        records.append(row)
                        values = [f"{key}: {value}" for key, value in row.items() if value not in (None, "")]
                        text_parts.append(f"Sheet {sheet.title} Row {index + 1}\n" + "\n".join(values))
                workbook.close()
            except Exception as exc:
                return {
                    "file_name": file_path.name,
                    "file_type": suffix.lstrip("."),
                    "content": f"[Spreadsheet parse error: {exc}]",
                    "metadata": {"error": str(exc)},
                }

        return {
            "file_name": file_path.name,
            "file_type": suffix.lstrip("."),
            "content": "\n\n".join(text_parts) or "[Spreadsheet contains no readable rows.]",
            "metadata": {
                "rows": len(records),
                "size_bytes": file_path.stat().st_size,
                "records": records,
            },
        }
